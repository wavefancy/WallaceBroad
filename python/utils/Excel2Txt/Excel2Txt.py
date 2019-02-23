#!/usr/bin/env python3

"""

    Convert Excel file (xls, xlsx) file to txt file.

    @Author: wavefancy@gmail.com

    Usage:
        Excel2Txt.py -f filename [-i index]
        Excel2Txt.py -h | --help | -v | --version

    Notes:
        1. Read content from -f file, and output results to stdout.
        2. Line index start from 1.

    Options:
        -f filename   Input file name of excel file (.xls|.xlsx).
        -i index      Index for worksheet, default 1. Index starts from 1.
        -h --help     Show this screen.
        -v --version  Show version.

"""
import sys
from docopt import docopt
from signal import signal, SIGPIPE, SIG_DFL
signal(SIGPIPE,SIG_DFL)

class P(object):
    delimeter = '\t'

def get_all_sheets(excel_file):
    '''Get all sheets name from an excel file.'''
    sheets = []
    workbook = load_workbook(excel_file,read_only=True,data_only=True)
    all_worksheets = workbook.get_sheet_names()
    for worksheet_name in all_worksheets:
        sheets.append(worksheet_name)
    return sheets

def output_from_excel(excel_file, worksheet_name):
    '''Out put content for a sheet'''
    workbook = load_workbook(excel_file,read_only=True,data_only=True)

    try:
        worksheet = workbook.get_sheet_by_name(worksheet_name)

        for row in worksheet.iter_rows():
            lrow = []
            for cell in row:
                lrow.append(str(cell.value))

            # out = [x for x in lrow if x]
            sys.stdout.write('%s\n'%(P.delimeter.join(lrow)))

    except KeyError:
        sys.stderr.write("Could not find " + worksheet_name)
        sys.exit(-1)

if __name__ == '__main__':
    args = docopt(__doc__, version='1.0')
#     print(args)

    # if(args['--format']):
    #     ShowFormat()
    #     sys.exit(-1)

    sheet_index = 0;
    if args['-i']:
        sheet_index = int(args['-i']) -1;

    filename = args['-f']

    from openpyxl import load_workbook
    sheets = get_all_sheets(filename)
    sheet = sheets[sheet_index]

    output_from_excel(filename, sheet)

sys.stdout.flush()
sys.stdout.close()
sys.stderr.flush()
sys.stderr.close()
